from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font
from domain.cfdi import CFDIRecord


HEADERS = list(CFDIRecord().__dict__.keys())


def write_sheet(ws, records: List[CFDIRecord]):
    # encabezados
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    row_idx = 2

    for record in records:
        for col_idx, value in enumerate(record.to_row(), 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        row_idx += 1


def build_excel_report(data):
    wb = Workbook()

    ws_ing = wb.active
    ws_ing.title = "INGRESOS"

    ws_egr = wb.create_sheet("EGRESOS")

    write_sheet(ws_ing, data["ingresos"])
    write_sheet(ws_egr, data["egresos"])

    return wb