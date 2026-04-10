# reporte_cfdi.py
from __future__ import annotations

import argparse
import os
import xml.etree.ElementTree as ET
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Dict, Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# =========================
# Configuración
# =========================

NAMESPACES = {
    "cfdi": "http://www.sat.gob.mx/cfd/4",
    "tfd": "http://www.sat.gob.mx/TimbreFiscalDigital",
}

OUTPUT_HEADERS = [
    "Version CFDI", "Mes", "Fecha Emision", "Tipo De Comprobante",
    "RFC Receptor", "Nombre Receptor", "RFC Emisor", "Nombre Emisor",
    "Serie", "Folio", "UUID", "Metodo Pago", "Imp Local Trasladado",
    "Forma Pago", "Moneda", "Tipo Cambio", "SubTotal", "IVA Trasladado8",
    "IVA Trasladado16", "IEPS Trasladado", "IEPS Trasladado Cuota",
    "Total Impuestos Retenidos", "IVA Retenido", "ISR Retenido", "Descuento",
    "Total", "Estatus", "CFDI Relacionados", "Fecha Cancelacion",
    "Estatus Cancelacion", "Imp Local Retenido", "Conceptos",
    "Regimen Fiscal Emisor", "Residencia Fiscal Receptor",
    "Imp Local Importe Traslado", "Tiene IVA Exento", "Base Tasa Cero",
    "IEPS Retenido Cuota", "Domicilio Fiscal Receptor"
]


# =========================
# Modelo de datos
# =========================

@dataclass
class CFDIRecord:
    version_cfdi: str = ""
    mes: Optional[int] = None
    fecha_emision: str = ""
    tipo_comprobante: str = ""
    rfc_receptor: str = ""
    nombre_receptor: str = ""
    rfc_emisor: str = ""
    nombre_emisor: str = ""
    serie: str = ""
    folio: str = ""
    uuid: str = ""
    metodo_pago: str = ""
    imp_local_trasladado: Optional[float] = None
    forma_pago: str = ""
    moneda: str = ""
    tipo_cambio: Optional[float] = None
    subtotal: Optional[float] = None
    iva_trasladado8: Optional[float] = None
    iva_trasladado16: Optional[float] = None
    ieps_trasladado: Optional[float] = None
    ieps_trasladado_cuota: Optional[float] = None
    total_impuestos_retenidos: Optional[float] = None
    iva_retenido: Optional[float] = None
    isr_retenido: Optional[float] = None
    descuento: Optional[float] = None
    total: Optional[float] = None
    estatus: str = ""
    cfdi_relacionados: str = ""
    fecha_cancelacion: str = ""
    estatus_cancelacion: str = ""
    imp_local_retenido: Optional[float] = None
    conceptos: str = ""
    regimen_fiscal_emisor: str = ""
    residencia_fiscal_receptor: str = ""
    imp_local_importe_traslado: Optional[float] = None
    tiene_iva_exento: str = ""
    base_tasa_cero: Optional[float] = None
    ieps_retenido_cuota: Optional[float] = None
    domicilio_fiscal_receptor: str = ""

    def to_row(self) -> List[Any]:
        return [
            self.version_cfdi,
            self.mes,
            self.fecha_emision,
            self.tipo_comprobante,
            self.rfc_receptor,
            self.nombre_receptor,
            self.rfc_emisor,
            self.nombre_emisor,
            self.serie,
            self.folio,
            self.uuid,
            self.metodo_pago,
            self.imp_local_trasladado,
            self.forma_pago,
            self.moneda,
            self.tipo_cambio,
            self.subtotal,
            self.iva_trasladado8,
            self.iva_trasladado16,
            self.ieps_trasladado,
            self.ieps_trasladado_cuota,
            self.total_impuestos_retenidos,
            self.iva_retenido,
            self.isr_retenido,
            self.descuento,
            self.total,
            self.estatus,
            self.cfdi_relacionados,
            self.fecha_cancelacion,
            self.estatus_cancelacion,
            self.imp_local_retenido,
            self.conceptos,
            self.regimen_fiscal_emisor,
            self.residencia_fiscal_receptor,
            self.imp_local_importe_traslado,
            self.tiene_iva_exento,
            self.base_tasa_cero,
            self.ieps_retenido_cuota,
            self.domicilio_fiscal_receptor,
        ]


# =========================
# Utilidades
# =========================

def safe_float(value: Optional[str]) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except ValueError:
        return None


def safe_str(value: Optional[str]) -> str:
    return "" if value is None else str(value)


def get_attr(node: ET.Element, attr: str, default: str = "") -> str:
    return node.attrib.get(attr, default)


def parse_fecha_iso(fecha: str) -> str:
    # Deja la fecha como texto ISO limpio
    if not fecha:
        return ""
    try:
        return datetime.fromisoformat(fecha).isoformat(sep="T", timespec="seconds")
    except ValueError:
        return fecha


def month_from_fecha(fecha: str) -> Optional[int]:
    if not fecha:
        return None
    try:
        return datetime.fromisoformat(fecha).month
    except ValueError:
        return None


# =========================
# Extracción XML
# =========================

def parse_cfdi_xml(xml_path: Path) -> Optional[CFDIRecord]:
    try:
        tree = ET.parse(xml_path)
        root = tree.getroot()
    except ET.ParseError as e:
        print(f"[WARN] XML inválido: {xml_path.name} -> {e}")
        return None

    version = root.attrib.get("Version", root.attrib.get("version", ""))
    fecha = parse_fecha_iso(root.attrib.get("Fecha", ""))
    tipo = root.attrib.get("TipoDeComprobante", "")
    serie = root.attrib.get("Serie", "")
    folio = root.attrib.get("Folio", "")
    moneda = root.attrib.get("Moneda", "")
    metodo_pago = root.attrib.get("MetodoPago", "")
    forma_pago = root.attrib.get("FormaPago", "")
    subtotal = safe_float(root.attrib.get("SubTotal"))
    total = safe_float(root.attrib.get("Total"))
    descuento = safe_float(root.attrib.get("Descuento"))
    tipo_cambio = safe_float(root.attrib.get("TipoCambio"))

    emisor = root.find("cfdi:Emisor", NAMESPACES)
    receptor = root.find("cfdi:Receptor", NAMESPACES)
    conceptos = root.findall("cfdi:Conceptos/cfdi:Concepto", NAMESPACES)
    impuestos = root.find("cfdi:Impuestos", NAMESPACES)
    complemento = root.find("cfdi:Complemento", NAMESPACES)
    tfd = None
    if complemento is not None:
        tfd = complemento.find("tfd:TimbreFiscalDigital", NAMESPACES)

    uuid = ""
    if tfd is not None:
        uuid = tfd.attrib.get("UUID", "")

    rfc_emisor = emisor.attrib.get("Rfc", "") if emisor is not None else ""
    nombre_emisor = emisor.attrib.get("Nombre", "") if emisor is not None else ""
    regimen_fiscal_emisor = emisor.attrib.get("RegimenFiscal", "") if emisor is not None else ""

    rfc_receptor = receptor.attrib.get("Rfc", "") if receptor is not None else ""
    nombre_receptor = receptor.attrib.get("Nombre", "") if receptor is not None else ""
    residencia_fiscal_receptor = receptor.attrib.get("ResidenciaFiscal", "") if receptor is not None else ""
    domicilio_fiscal_receptor = receptor.attrib.get("DomicilioFiscalReceptor", "") if receptor is not None else ""

    iva16 = 0.0
    iva8 = 0.0
    ieps = 0.0
    ieps_cuota = 0.0
    iva_retenido = 0.0
    isr_retenido = 0.0
    imp_local_trasladado = 0.0
    imp_local_retenido = 0.0
    total_impuestos_retenidos = 0.0
    base_tasa_cero = 0.0
    tiene_iva_exento = False
    imp_local_importe_traslado = 0.0

    # Impuestos globales
    if impuestos is not None:
        trasladados = impuestos.find("cfdi:Traslados", NAMESPACES)
        retenidos = impuestos.find("cfdi:Retenciones", NAMESPACES)

        if trasladados is not None:
            for tr in trasladados.findall("cfdi:Traslado", NAMESPACES):
                impuesto = tr.attrib.get("Impuesto", "")
                tasa = tr.attrib.get("TasaOCuota", "")
                base = safe_float(tr.attrib.get("Base")) or 0.0
                importe = safe_float(tr.attrib.get("Importe")) or 0.0

                if impuesto == "002":  # IVA
                    if tasa == "0.160000":
                        iva16 += importe
                    elif tasa == "0.080000":
                        iva8 += importe
                    elif tasa == "0.000000":
                        base_tasa_cero += base
                    else:
                        # si no cuadra exacto, lo agregamos al de 16 por defecto
                        iva16 += importe
                elif impuesto == "003":  # IEPS
                    ieps += importe

                # impuestos locales
                if tr.attrib.get("ImpuestoLocal") or tr.attrib.get("TipoFactor") == "Cuota":
                    imp_local_trasladado += importe
                    imp_local_importe_traslado += importe
                    if tr.attrib.get("TipoFactor") == "Cuota":
                        ieps_cuota += importe

        if retenidos is not None:
            for ret in retenidos.findall("cfdi:Retencion", NAMESPACES):
                impuesto = ret.attrib.get("Impuesto", "")
                importe = safe_float(ret.attrib.get("Importe")) or 0.0
                total_impuestos_retenidos += importe

                if impuesto == "002":
                    iva_retenido += importe
                elif impuesto == "001":
                    isr_retenido += importe
                elif impuesto == "003":
                    ieps_cuota += importe  # si quieres separarlo más adelante, cámbialo aquí

    # Conceptos: texto compacto
    conceptos_texto = []
    for c in conceptos:
        desc = c.attrib.get("Descripcion", "")
        if desc:
            conceptos_texto.append(desc)
        # detectar IVA exento en conceptos
        imp_con = c.find("cfdi:Impuestos", NAMESPACES)
        if imp_con is not None:
            tras = imp_con.find("cfdi:Traslados", NAMESPACES)
            if tras is not None:
                for tr in tras.findall("cfdi:Traslado", NAMESPACES):
                    if tr.attrib.get("Impuesto") == "002" and tr.attrib.get("TasaOCuota") == "0.000000":
                        tiene_iva_exento = True

    # Clasificación para estatus base
    estatus = "Vigente"  # en XML normalmente no viene cancelado; se deja como base
    tipo_comprobante_legible = {
        "I": "I-Ingreso",
        "E": "E-Egreso",
        "P": "P-Pago",
        "N": "N-Nómina",
        "T": "T-Traslado",
    }.get(tipo, tipo)

    record = CFDIRecord(
        version_cfdi=version,
        mes=month_from_fecha(fecha),
        fecha_emision=fecha,
        tipo_comprobante=tipo_comprobante_legible,
        rfc_receptor=rfc_receptor,
        nombre_receptor=nombre_receptor,
        rfc_emisor=rfc_emisor,
        nombre_emisor=nombre_emisor,
        serie=serie,
        folio=folio,
        uuid=uuid,
        metodo_pago=metodo_pago,
        imp_local_trasladado=imp_local_trasladado if imp_local_trasladado else None,
        forma_pago=forma_pago,
        moneda=moneda,
        tipo_cambio=tipo_cambio,
        subtotal=subtotal,
        iva_trasladado8=iva8 if iva8 else None,
        iva_trasladado16=iva16 if iva16 else None,
        ieps_trasladado=ieps if ieps else None,
        ieps_trasladado_cuota=ieps_cuota if ieps_cuota else None,
        total_impuestos_retenidos=total_impuestos_retenidos if total_impuestos_retenidos else None,
        iva_retenido=iva_retenido if iva_retenido else None,
        isr_retenido=isr_retenido if isr_retenido else None,
        descuento=descuento,
        total=total,
        estatus=estatus,
        cfdi_relacionados="",
        fecha_cancelacion="",
        estatus_cancelacion="",
        imp_local_retenido=imp_local_retenido if imp_local_retenido else None,
        conceptos=" | ".join(conceptos_texto[:5]),
        regimen_fiscal_emisor=regimen_fiscal_emisor,
        residencia_fiscal_receptor=residencia_fiscal_receptor,
        imp_local_importe_traslado=imp_local_importe_traslado if imp_local_importe_traslado else None,
        tiene_iva_exento="Sí" if tiene_iva_exento else "No",
        base_tasa_cero=base_tasa_cero if base_tasa_cero else None,
        ieps_retenido_cuota=ieps_cuota if ieps_cuota else None,
        domicilio_fiscal_receptor=domicilio_fiscal_receptor,
    )
    return record


# =========================
# Excel
# =========================

def build_workbook(records: List[CFDIRecord], template_path: Optional[Path] = None) -> Workbook:
    if template_path and template_path.exists():
        wb = load_workbook(template_path)
    else:
        wb = Workbook()
        # limpiar hoja por defecto
        if wb.active.title == "Sheet":
            wb.remove(wb.active)

    # asegurar hojas
    for sheet_name in ["INGRESOS", "EGRESOS"]:
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)

    ingresos = [r for r in records if r.tipo_comprobante.startswith("I-")]
    egresos = [r for r in records if r.tipo_comprobante.startswith("E-")]

    write_sheet(wb["INGRESOS"], ingresos)
    write_sheet(wb["EGRESOS"], egresos)

    return wb


def write_sheet(ws, records: List[CFDIRecord]) -> None:
    # encabezados
    for col_idx, header in enumerate(OUTPUT_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # datos
    for row_idx, record in enumerate(records, start=2):
        row = record.to_row()
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # formato básico
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(OUTPUT_HEADERS))}{max(1, len(records) + 1)}"

    widths = {
        1: 14, 2: 8, 3: 22, 4: 16, 5: 18, 6: 28, 7: 18, 8: 28,
        9: 12, 10: 12, 11: 38, 12: 20, 13: 15, 14: 20, 15: 10, 16: 12,
        17: 14, 18: 14, 19: 14, 20: 14, 21: 16, 22: 18, 23: 12, 24: 12,
        25: 12, 26: 14, 27: 12, 28: 26, 29: 16, 30: 16, 31: 14, 32: 30,
        33: 18, 34: 18, 35: 16, 36: 14, 37: 14, 38: 14, 39: 18,
    }
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width

    # formato de fecha / texto general
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.column in {2}:
                cell.alignment = Alignment(horizontal="center")
            elif cell.column in {3, 29}:
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(vertical="top", wrap_text=True)


# =========================
# Main
# =========================

def collect_xml_files(input_dir: Path) -> List[Path]:
    return sorted([p for p in input_dir.rglob("*.xml") if p.is_file()])


def main():
    parser = argparse.ArgumentParser(description="Leer XML CFDI y generar reporte Excel.")
    parser.add_argument("--input", required=True, help="Carpeta con XML")
    parser.add_argument("--output", required=True, help="Archivo Excel de salida .xlsx")
    parser.add_argument("--template", default="", help="Plantilla Excel opcional")
    parser.add_argument("--rfc", required=True, help="Definir el rfc del cliente")

    args = parser.parse_args()

    input_dir = Path(args.input)
    output_path = Path(args.output)
    template_path = Path(args.template) if args.template else None

    if not input_dir.exists():
        raise FileNotFoundError(f"No existe la carpeta de entrada: {input_dir}")

    xml_files = collect_xml_files(input_dir)
    if not xml_files:
        print("[INFO] No se encontraron XML.")
        return

    records: List[CFDIRecord] = []
    for xml_file in xml_files:
        record = parse_cfdi_xml(xml_file)
        if record is not None:
            records.append(record)

    print(records[0])

    emitidos = [r for r in records if r.rfc_emisor == args.rfc]
    recibidos = [r for r in records if r.rfc_receptor == args.rfc]

    wb = build_workbook(records, template_path=template_path)
    wb.save(output_path)

    print(f"[OK] XML procesados: {len(xml_files)}")
    print(f"[OK] Registros extraídos: {len(records)}")
    print(f"[OK] Archivo generado: {output_path}")


if __name__ == "__main__":
    main()